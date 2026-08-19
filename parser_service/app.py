# parser_service/app.py
# Robust FastAPI Parser for Amazon DSP KPI PDFs with ranking & status bucket
from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any, Tuple
import io, re
import pdfplumber
import pandas as pd

app = FastAPI(title="Amazon DSP KPI Parser")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        # Local development
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:4000",
        "http://127.0.0.1:4000",
        "http://localhost:5000",
        "http://127.0.0.1:5000",
        "http://localhost:8080",
        "http://127.0.0.1:8080",
        # Production (Firebase Hosting + Custom Domain) — codriver-eu only.
        "https://dsp-codriver.de",
        "https://www.dsp-codriver.de",
        "https://codriver-eu.web.app",
        "https://codriver-eu.firebaseapp.com",
    ],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ===============================
# MODELS
# ===============================
class DriverRow(BaseModel):
    transporter_id: str = Field(alias="Transporter ID")
    Delivered: Optional[float] = None
    DCR: Optional[float] = None
    POD: Optional[float] = None
    CC: Optional[float] = None
    CE: Optional[float] = None
    LoR_DPMO: Optional[float] = Field(None, alias="LoR DPMO")
    DNR_DPMO: Optional[float] = Field(None, alias="DNR DPMO")
    CDF_DPMO: Optional[float] = Field(None, alias="CDF DPMO")

    # Computed per Albert
    POD_Score: Optional[float] = None
    CC_Score: Optional[float] = None
    DCR_Score: Optional[float] = None
    CE_Score: Optional[float] = None
    LoR_Score: Optional[float] = None
    DNR_Score: Optional[float] = None
    CDF_Score: Optional[float] = None
    FinalScore: Optional[float] = None  # "TOTAL"

    # Added for UI
    rank: Optional[int] = None
    statusBucket: Optional[str] = None

    # POD Quality (driver-level counts)
    POD_Q_Opportunities: Optional[float] = None
    POD_Q_Success: Optional[float] = None
    POD_Q_Bypass: Optional[float] = None
    POD_Q_Rejects: Optional[float] = None
    POD_Q_BlurryPhoto: Optional[float] = None
    POD_Q_PhotoTooDark: Optional[float] = None
    POD_Q_NoPackageDetected: Optional[float] = None
    POD_Q_PackageInCar: Optional[float] = None
    POD_Q_PackageTooClose: Optional[float] = None

    # Concessions (driver-level)
    CONC_TotalDelivered: Optional[float] = None
    CONC_TotalDnr: Optional[float] = None
    CONC_DnrDpmo4w: Optional[float] = None
    CONC_DnrCount_W1: Optional[float] = None
    CONC_DnrCount_W2: Optional[float] = None
    CONC_DnrCount_W3: Optional[float] = None
    CONC_DnrCount_W4: Optional[float] = None
    CONC_DnrDpmo_W1: Optional[float] = None
    CONC_DnrDpmo_W2: Optional[float] = None
    CONC_DnrDpmo_W3: Optional[float] = None
    CONC_DnrDpmo_W4: Optional[float] = None
    CONC_Focus_AttendedDnr: Optional[float] = None
    CONC_Focus_PhotoOnDelivery: Optional[float] = None
    CONC_Focus_SuccessfulContact: Optional[float] = None
    CONC_Focus_Delivered25m: Optional[float] = None
    CONC_Focus_FalseScan: Optional[float] = None
    CONC_Focus_Mailbox: Optional[float] = None
    CONC_Focus_DeliveredOtp: Optional[float] = None

    # Customer Delivery Feedback (CDF, driver-level)
    CDF_TotalFeedback: Optional[float] = None
    CDF_NegativeFeedback: Optional[float] = None
    CDF_L1_NeverReceived: Optional[float] = None
    CDF_L1_DriverMishandled: Optional[float] = None
    CDF_L1_NotDeliveredToPreferredLocation: Optional[float] = None
    CDF_L1_DamagedPackage: Optional[float] = None
    CDF_L1_DeliveryWasLate: Optional[float] = None
    CDF_L1_BadDeliveryExperience: Optional[float] = None
    CDF_L1_DriverWasUnprofessional: Optional[float] = None
    CDF_L1_Other: Optional[float] = None
    # Event-level details (raw per-event rows for the detail UI)
    CDF_Events: Optional[List[Dict[str, Any]]] = None

    # DWC / IADC (driver-level, from the weekly HTML report)
    DWC_DwcPct: Optional[float] = None
    DWC_IadcPct: Optional[float] = None
    # Per-category miss counts, e.g. { "Photo Defect": 3, "Contact Miss": 5 }.
    DWC_Misses: Optional[Dict[str, int]] = None

class ParserSummary(BaseModel):
    overallScore: Optional[float] = None
    overallStatus: Optional[str] = None
    reliabilityScore: Optional[float] = None
    rankAtStation: Optional[int] = None
    stationCount: Optional[int] = None
    rankDeltaWoW: Optional[int] = None
    weekText: Optional[str] = None
    weekNumber: Optional[int] = None
    year: Optional[int] = None
    stationCode: Optional[str] = None

    reliabilityNextDay: Optional[float] = None
    reliabilitySameDay: Optional[float] = None

    podQualitySummary: Optional[Dict[str, Any]] = None
    podQualityRejects: Optional[Dict[str, Any]] = None
    complianceAndSafety: Optional[Dict[str, Any]] = None
    deliveryQualitySwc: Optional[Dict[str, Any]] = None

    concessionsSummary: Optional[Dict[str, Any]] = None
    concessionsFocusBuckets: Optional[Dict[str, Any]] = None

    cdfSummary: Optional[Dict[str, Any]] = None
    cdfL1Distribution: Optional[Dict[str, Any]] = None

    dwcSummary: Optional[Dict[str, Any]] = None

class ParserResponse(BaseModel):
    count: int
    drivers: List[DriverRow]
    summary: Optional[ParserSummary] = None
    # Cortex fleet import — populated by /parse when the uploaded file
    # is a VehiclesData XLSX. Vehicles are pass-through dicts (no strict
    # schema yet) so the Flutter dedup logic can match on FIN.
    vehicles: Optional[List[Dict[str, Any]]] = None

    # Amazon DSP invoice / Gutschrift — populated when the uploaded file is
    # a weekly credit note. Pass-through dict: { invoiceNumber, week,
    # rangeStart, rangeEnd, total, lines: [{date, serviceType, unitPrice,
    # quantity, total}, ...] }. Used by the Payment Check reconciliation.
    invoice: Optional[Dict[str, Any]] = None

# ===============================
# HELPERS
# ===============================
NBSP = "\u00A0"

def clean_str(x: Any) -> str:
    if x is None:
        return ""
    s = str(x)
    s = (
        s.replace("\u00A0", " ")
         .replace("\u2009", " ")
         .replace("\u202F", " ")
         .replace("\u00AD", "")
         .replace("\n", " ")
         .replace("\r", " ")
    )
    return re.sub(r"\s+", " ", s).strip()

def keyize(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())

def to_num(x: Any) -> Optional[float]:
    if x is None:
        return None
    s = clean_str(x)
    if not s or s == "-":
        return None
    s = s.replace("%", "").replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s)
    except:
        return None

def to_percent(x: Any) -> Optional[float]:
    if x is None:
        return None
    s = clean_str(x)
    if not s or s == "-":
        return None
    if s.endswith("%"):
        return to_num(s[:-1])
    n = to_num(s)
    if n is None:
        return None
    return n * 100.0 if n < 1.0 else n

def clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))

def normalize_status_label(raw: Any) -> str:
    s = clean_str(raw)
    if not s:
        return ""
    s = re.sub(r"[\s\-]+", "_", s.upper())
    s = re.sub(r"[^A-Z0-9_]", "", s)
    return s.strip("_")

def status_bucket(
    final_score: Optional[float],
    week_number: Optional[int] = None,
    year: Optional[int] = None,
) -> str:
    if final_score is None:
        return "Unknown"
    s = float(final_score)
    if s >= 93.0:
        return "FANTASTIC_PLUS"
    if s >= 86.0:
        return "FANTASTIC"
    if s >= 70:
        return "GREAT"
    if s >= 50:
        return "FAIR"
    if s >= 0:
        return "POOR"
    return "POOR"

def _extract_metric_token(
    text: str,
    label: str,
    next_labels: List[str],
) -> Optional[str]:
    lookahead = ""
    if next_labels:
        joined = "|".join(re.escape(v) for v in next_labels)
        lookahead = rf"(?=\s+(?:{joined})\b|$)"
    pattern = re.compile(rf"{re.escape(label)}\s+(.+?){lookahead}", re.IGNORECASE)
    match = pattern.search(text)
    if not match:
        return None
    return clean_str(match.group(1))

def _extract_tokens_by_labels(
    text: str,
    labels: List[str],
) -> Dict[str, str]:
    found: List[tuple[str, int]] = []
    cursor = 0
    lower_text = text.lower()

    for label in labels:
        idx = lower_text.find(label.lower(), cursor)
        if idx < 0:
            continue
        found.append((label, idx))
        cursor = idx + len(label)

    tokens: Dict[str, str] = {}
    for i, (label, start) in enumerate(found):
        value_start = start + len(label)
        value_end = found[i + 1][1] if i + 1 < len(found) else len(text)
        tokens[label] = clean_str(text[value_start:value_end])
    return tokens

def _extract_section_overall_status(
    text: str,
    section_label: str,
    stop_labels: List[str],
) -> str:
    lower_text = text.lower()
    start = lower_text.find(section_label.lower())
    if start < 0:
        return ""

    start += len(section_label)
    end = len(text)
    for stop_label in stop_labels:
        idx = lower_text.find(stop_label.lower(), start)
        if idx >= 0:
            end = min(end, idx)

    snippet = clean_str(text[start:end])
    if not snippet:
        return ""

    match = re.search(r"\b(Fantastic Plus|Fantastic|Great|Fair|Poor)\b", snippet, re.IGNORECASE)
    return normalize_status_label(match.group(1)) if match else ""

def _parse_metric_value(token: Optional[str]) -> Optional[Dict[str, Any]]:
    s = clean_str(token)
    if not s:
        return None
    s = re.sub(r"^[^A-Za-z0-9]+", "", s)
    s = re.sub(r"\s+", " ", s).strip()

    value_part = s
    status_part = ""
    if "|" in s:
        left, right = s.split("|", 1)
        value_part = clean_str(left)
        status_part = clean_str(right)

    parsed_num = to_num(value_part)
    value: Any = parsed_num if parsed_num is not None else value_part
    out: Dict[str, Any] = {"value": value}
    if status_part:
        out["status"] = normalize_status_label(status_part)
    return out

def _parse_scorecard_summary_sections(pdf: pdfplumber.PDF) -> Dict[str, Any]:
    page_text = ""
    for page in pdf.pages[:3]:
        text = page.extract_text() or ""
        if "Compliance and Safety" in text and "Delivery Quality & SWC" in text:
            page_text = text
            break
    if not page_text:
        return {}

    flat = clean_str(page_text)
    result: Dict[str, Any] = {}

    safety_overall = _extract_section_overall_status(
        flat,
        "Compliance and Safety",
        [
            "Safe Driving Metric (FICO)",
            "Safety",
            "Safety Compliance",
            "Delivery Quality & SWC:",
        ],
    )
    delivery_overall = _extract_section_overall_status(
        flat,
        "Delivery Quality & SWC:",
        [
            "Customer delivery Experience",
            "Customer escalation DPMO",
            "Quality",
            "Metrics highlighted in red",
        ],
    )

    safety_labels = [
        ("Safe Driving Metric (FICO)", "fico"),
        ("Vehicle Audit (VSA) Compliance", "vehicleAuditCompliance"),
        ("Speeding Event Rate (Per 100 Trips)", "speedingEventRatePer100Trips"),
        ("Breach of Contract (BOC)", "breachOfContract"),
        ("Mentor Adoption Rate", "mentorAdoptionRate"),
        ("Working Hours Compliance (WHC)", "workingHoursCompliance"),
        ("Comprehensive Audit Score (CAS)", "comprehensiveAuditScore"),
    ]
    delivery_labels = [
        ("Customer escalation DPMO", "customerEscalationDpmo"),
        ("Delivery Completion Rate(DCR)", "deliveryCompletionRate"),
        ("Customer Delivery Feedback", "customerDeliveryFeedback"),
        ("Delivered Not Received(DNR DPMO)", "dnrDpmo"),
        ("Lost on Road (LoR) DPMO", "lorDpmo"),
        ("Delivery Success Conditions (DSC DPMO)", "dscDpmo"),
        ("Photo-On-Delivery", "photoOnDelivery"),
        ("Contact Compliance", "contactCompliance"),
    ]
    safety_tokens = _extract_tokens_by_labels(
        flat,
        [label for label, _ in safety_labels] + ["Delivery Quality & SWC:"],
    )
    delivery_tokens = _extract_tokens_by_labels(
        flat,
        [
            "Customer escalation DPMO",
            "Delivery Completion Rate(DCR)",
            "Customer Delivery Feedback",
            "Delivered Not Received(DNR DPMO)",
            "Lost on Road (LoR) DPMO",
            "Standard Work Compliance",
            "Delivery Success Conditions (DSC DPMO)",
            "Photo-On-Delivery",
            "Contact Compliance",
            "Metrics highlighted in red",
        ],
    )

    compliance_and_safety = {
        "overallStatus": safety_overall,
        "safety": {},
        "compliance": {},
    }
    for label, key in safety_labels:
        token = safety_tokens.get(label)
        metric = _parse_metric_value(token)
        if metric is None:
            continue
        if key in {"fico", "speedingEventRatePer100Trips", "mentorAdoptionRate"}:
            compliance_and_safety["safety"][key] = metric
        else:
            compliance_and_safety["compliance"][key] = metric

    if compliance_and_safety["overallStatus"] or compliance_and_safety["safety"] or compliance_and_safety["compliance"]:
        result["complianceAndSafety"] = compliance_and_safety

    delivery_quality_swc = {
        "overallStatus": delivery_overall,
        "customerDeliveryExperience": {},
        "quality": {},
        "standardWorkCompliance": {},
    }
    for label, key in delivery_labels:
        token = delivery_tokens.get(label)
        metric = _parse_metric_value(token)
        if metric is None:
            continue
        if key in {"customerEscalationDpmo", "customerDeliveryFeedback"}:
            delivery_quality_swc["customerDeliveryExperience"][key] = metric
        elif key in {"deliveryCompletionRate", "dnrDpmo", "lorDpmo"}:
            delivery_quality_swc["quality"][key] = metric
        else:
            delivery_quality_swc["standardWorkCompliance"][key] = metric

    if delivery_quality_swc["overallStatus"] or delivery_quality_swc["customerDeliveryExperience"] or delivery_quality_swc["quality"] or delivery_quality_swc["standardWorkCompliance"]:
        result["deliveryQualitySwc"] = delivery_quality_swc

    return result

# ===============================
# KPI FORMULAS (Albert’s rules)
# ===============================
def compute_scores(
    row: Dict[str, Any],
    week_number: Optional[int] = None,
    year: Optional[int] = None,   # <-- ADDED (only to decide FinalScore regime)
) -> Dict[str, Optional[float]]:
    # --- parse raw values (unchanged) ---
    pod = to_percent(row.get("POD"))
    cc  = to_percent(row.get("CC"))
    dcr = to_percent(row.get("DCR"))
    ce  = to_num(row.get("CE"))
    lor = to_num(row.get("LoR DPMO"))
    dnr = to_num(row.get("DNR DPMO"))
    cdf = to_num(row.get("CDF DPMO"))
    delivered = to_num(row.get("Delivered"))

    ce_val  = 0.0 if ce  is None else ce
    lor_val = 0.0 if lor is None else lor
    dnr_val = 0.0 if dnr is None else dnr
    cdf_val = 0.0 if cdf is None else cdf

    # --- per-metric scores (unchanged) ---
    POD_Score = None if pod is None else clamp(pod, 0, 100)
    CC_Score  = None if cc  is None else clamp(cc,  0, 100)
    DCR_Score = None if dcr is None else clamp(dcr, 0, 100)
    CE_Score  = None if ce  is None else clamp(100 - 50.0 * ce, 50, 100)

    CE_Score  = clamp(100.0 - ce_val * 50.0, 50.0, 100.0)
    LoR_Score = clamp(100.0 - (lor_val / 1200.0) * 30.0, 0.0, 100.0)
    DNR_Score = clamp(100.0 - (dnr_val / 1200.0) * 30.0, 0.0, 100.0)

    # UPDATED: CDF score mode is decided strictly by header presence (not by value).
    cdf_mode = row.get("_cdf_mode")  # "PCT" | "DPMO" | None
    if cdf_mode == "PCT":
        cdf_pct = to_percent(row.get("CDF"))
        CDF_Score = clamp(cdf_pct, 0.0, 100.0)
    elif cdf_mode == "DPMO":
        cdf_dpmo = to_num(row.get("CDF DPMO"))
        cdf_dpmo_val = 0.0 if cdf_dpmo is None else cdf_dpmo
        CDF_Score = clamp(100.33 - 0.01333 * float(cdf_dpmo_val), 0.0, 100.0)
    else:
        CDF_Score = None

    CE_Score  = round(CE_Score, 2)
    LoR_Score = round(LoR_Score, 2)
    DNR_Score = round(DNR_Score, 2)
    CDF_Score = None if CDF_Score is None else round(CDF_Score, 2)

    def z(x: Optional[float]) -> float:
        return 0.0 if x is None else float(x)

    FinalScore: Optional[float] = None

    def dsc_penalty_rate(dpmo: float) -> float:
        if dpmo <= 0:
            return 0.0
        if dpmo < 600:
            return 0.05
        if dpmo <= 1000:
            return 0.08
        if dpmo <= 2500:
            return 0.20
        if dpmo <= 5000:
            return 0.35
        if dpmo <= 10000:
            return 0.50
        return 0.70

    def lor_penalty_rate(dpmo: float) -> float:
        if dpmo <= 0:
            return 0.0
        if dpmo <= 500:
            return 0.10
        if dpmo <= 1500:
            return 0.25
        if dpmo <= 4000:
            return 0.40
        return 0.80

    # Final score uses the KPI average as the base, then applies:
    #   1) a direct CE penalty
    #   2) a DSC/DNR DPMO percentage penalty based on manager thresholds
    #   3) a LoR DPMO percentage penalty based on manager thresholds
    vals = [v for v in [DCR_Score, POD_Score, CC_Score, CDF_Score] if v is not None]
    if vals:
        avg_base = sum(vals) / len(vals)
        ce_penalty = 90.0 * abs(ce_val)
        dsc_penalty = avg_base * dsc_penalty_rate(dnr_val)
        lor_penalty = avg_base * lor_penalty_rate(lor_val)
        FinalScore = avg_base - ce_penalty - dsc_penalty - lor_penalty

    return {
        "POD_Score": POD_Score, "CC_Score": CC_Score, "DCR_Score": DCR_Score,
        "CE_Score": CE_Score, "LoR_Score": LoR_Score, "DNR_Score": DNR_Score,
        "CDF_Score": CDF_Score, "FinalScore": round(FinalScore, 2)
    }

# ===============================
# HEADER NORMALIZATION
# ===============================
HEADER_MAP = {
    "transporter id": "Transporter ID",
    "zustellende-id": "Transporter ID",
    "associate id": "Transporter ID",
    "driver id": "Transporter ID",
    "dcr": "DCR",
    "pod": "POD",
    "cc": "CC",
    "ce": "CE",
    "lor dpmo": "LoR DPMO",
    "lor (dpmo)": "LoR DPMO",
    "dnr dpmo": "DNR DPMO",
    "dnr (dpmo)": "DNR DPMO",
    "dsc dpmo": "DNR DPMO",
    "DSC DPMO": "DNR DPMO",
    "dsc (dpmo)": "DNR DPMO",

    "cdf dpmo": "CDF DPMO",
    "cdf (dpmo)": "CDF DPMO",
    "cdfdpmo": "CDF DPMO",
    "cdfdpm0": "CDF DPMO",
    "cdf": "CDF",

    "delivered": "Delivered",
    "zugestellte pakete": "Delivered",
}

def norm_col(c: Any) -> str:
    raw = clean_str(c)
    k = keyize(raw)
    return HEADER_MAP.get(k, raw)

# ===============================
# PDF TABLE EXTRACTION
# ===============================
def get_col(df: pd.DataFrame, canonical: str) -> Optional[str]:
    want = keyize(canonical)
    for col in df.columns:
        if keyize(col) == want:
            return col
    return None

# -------------------------------------------------------------------
# FALLBACK: single-column ("columnless") driver tables
#
# Some Amazon Scorecard 3.0 PDFs are rendered without vertical rulings,
# so pdfplumber's extract_tables() returns every visual row as ONE cell:
#
#   header : 'S DSC LoR CDF\nTransporter ID Delivered DCR POD CC CE\nNo. DPMO DPMO DPMO'
#   data   : '1 A11QQXKOJRNZUJ 320 99.38% 3125 0 100.00% 100.00% 0 3125'
#
# The helpers below re-derive the real columns from the word geometry of
# the page (x-positions) and split the data lines with a strict regex.
# They are ONLY used when the regular column detection produced nothing.
# -------------------------------------------------------------------
_COLUMNLESS_DATA_RE = re.compile(r"^\s*(\d{1,4})\s+([A-Z0-9]{8,20})\s+(\S.*)$")
_COLUMNLESS_VALUE_RE = re.compile(r"^(?:-|[+-]?\d[\d.,]*\s?%?)$")


def _words_to_text_lines(words: List[Dict[str, Any]], tol: float = 1.6) -> List[Dict[str, Any]]:
    """Group pdfplumber words into visual text lines (by their `top` coordinate)."""
    lines: List[Dict[str, Any]] = []
    for w in sorted(words, key=lambda w: (float(w.get("top", 0.0)), float(w.get("x0", 0.0)))):
        top = float(w.get("top", 0.0))
        if lines and abs(float(lines[-1]["top"]) - top) <= tol:
            lines[-1]["words"].append(w)
        else:
            lines.append({"top": top, "words": [w]})
    for ln in lines:
        ln["words"].sort(key=lambda w: float(w.get("x0", 0.0)))
        ln["text"] = re.sub(r"\s+", " ", " ".join(str(w.get("text", "")) for w in ln["words"])).strip()
    return lines


def _group_words_into_columns(words: List[Dict[str, Any]], gap: float) -> List[str]:
    """Cluster header words horizontally; words of one column are joined top-to-bottom."""
    cols: List[Dict[str, Any]] = []
    for w in sorted(words, key=lambda w: float(w.get("x0", 0.0))):
        x0 = float(w.get("x0", 0.0)); x1 = float(w.get("x1", 0.0))
        if cols and x0 - cols[-1]["x1"] <= gap:
            cols[-1]["x1"] = max(cols[-1]["x1"], x1)
            cols[-1]["words"].append(w)
        else:
            cols.append({"x0": x0, "x1": x1, "words": [w]})
    out: List[str] = []
    for c in cols:
        c["words"].sort(key=lambda w: (float(w.get("top", 0.0)), float(w.get("x0", 0.0))))
        out.append(re.sub(r"\s+", " ", " ".join(str(w.get("text", "")) for w in c["words"])).strip())
    return out


def _reconstruct_columnless_header(page: Any, header_cell: Any, want_cols: int) -> Optional[List[str]]:
    """
    Rebuild the real column labels of a columnless table header.

    The header cell text contains the stacked header lines (e.g. 'S DSC LoR CDF' /
    'Transporter ID Delivered DCR POD CC CE' / 'No. DPMO DPMO DPMO'). We locate those
    lines in the page's word geometry and regroup the words by x-position, so the
    fragments become 'S No.', 'Transporter ID', ..., 'DSC DPMO', 'LoR DPMO', 'CDF DPMO'.

    The result is only accepted when it yields exactly `want_cols` columns (the number
    of values the data rows carry) - this self-validates the reconstruction.
    """
    raw_lines = [re.sub(r"\s+", " ", ln).strip() for ln in str(header_cell or "").split("\n")]
    raw_lines = [ln for ln in raw_lines if ln]
    if not raw_lines:
        return None
    try:
        words = page.extract_words() or []
    except Exception:
        return None
    if not words:
        return None

    plines = _words_to_text_lines(words)
    n = len(raw_lines)
    matched: Optional[List[Dict[str, Any]]] = None
    for i in range(len(plines) - n + 1):
        if all(plines[i + j]["text"] == raw_lines[j] for j in range(n)):
            matched = plines[i:i + n]
            break
    if matched is None:
        return None

    hwords = [w for ln in matched for w in ln["words"]]
    if not hwords:
        return None

    widths = []
    for w in hwords:
        t = str(w.get("text", ""))
        if t:
            widths.append((float(w.get("x1", 0.0)) - float(w.get("x0", 0.0))) / len(t))
    base = sorted(widths)[len(widths) // 2] if widths else 3.0

    # Try increasing word gaps; the first one that reproduces the data column count wins.
    for factor in (1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 6.0, 8.0):
        cols = _group_words_into_columns(hwords, max(3.0, base * factor))
        if len(cols) == want_cols and all(c for c in cols):
            return cols
    return None


def _parse_columnless_driver_table(
    page: Any,
    table: List[List[Any]],
) -> Optional[Tuple[List[str], List[List[str]]]]:
    """
    Detect + parse a driver table whose rows come back as a single cell each.
    Returns (header, data_rows) or None if the table is not such a driver table.
    """
    if not table or len(table) < 2:
        return None
    # every row must be a single-cell row (that is the whole point of this fallback)
    if any(len(r or []) != 1 for r in table):
        return None

    header_cell = table[0][0]
    if "transporter id" not in re.sub(r"\s+", " ", str(header_cell or "")).lower():
        return None

    parsed: List[List[str]] = []
    for r in table[1:]:
        cell = r[0]
        if cell is None:
            continue
        for line in str(cell).split("\n"):
            line = re.sub(r"[\u00A0\u2009\u202F]", " ", line)
            line = re.sub(r"\s+", " ", line).strip()
            if not line:
                continue
            m = _COLUMNLESS_DATA_RE.match(line)
            if not m:
                continue  # repeated header / footnote / non-driver line
            values = m.group(3).split(" ")
            if not all(_COLUMNLESS_VALUE_RE.match(v) for v in values):
                continue  # e.g. the working-hours table ('No'/'Yes' cells)
            parsed.append([m.group(1), m.group(2)] + values)

    if not parsed:
        return None

    counts: Dict[int, int] = {}
    for p in parsed:
        counts[len(p)] = counts.get(len(p), 0) + 1
    want_cols = max(counts.items(), key=lambda kv: (kv[1], kv[0]))[0]
    rows = [p for p in parsed if len(p) == want_cols]
    if not rows:
        return None

    header = _reconstruct_columnless_header(page, header_cell, want_cols)
    if not header:
        return None
    if not any(keyize(h) == "transporterid" for h in header):
        return None

    dropped = len(parsed) - len(rows)
    if dropped:
        print(f"Columnless driver table: dropped {dropped} row(s) with unexpected column count")
    print(f"Columnless driver table detected: {len(rows)} rows, header={header}")
    return header, rows


def extract_driver_rows(
    pdf: pdfplumber.PDF,
    week_number: Optional[int] = None,
    year: Optional[int] = None,   # <-- ADDED (only to pass into compute_scores)
) -> List[Dict[str, Any]]:
    def is_header_like(row: List[Any]) -> bool:
        if not row:
            return False
        texty = sum(1 for c in row if re.search(r"[A-Za-z]", clean_str(c)))
        return texty >= 3

    def looks_like_transporter_id(x: Any) -> bool:
        s = clean_str(x)
        return bool(re.match(r"^[A-Z0-9]{8,20}$", s))

    def coalesce_numbers(a, b):
        va = to_num(a)
        if va is not None:
            return va
        vb = to_num(b)
        return vb

    def merge_adjacent_cols_if_needed(df: pd.DataFrame, target_canonical: str) -> pd.DataFrame:
        want = keyize(target_canonical)
        cols = list(df.columns)
        for i in range(len(cols) - 1):
            k = keyize(str(cols[i]) + str(cols[i + 1]))
            if k == want:
                merged = []
                for _, row in df.iterrows():
                    merged.append(coalesce_numbers(row[cols[i]], row[cols[i + 1]]))
                df[target_canonical] = merged
                df = df.drop(columns=[cols[i], cols[i + 1]])
                return df
        return df

    def split_combo_headers(header: List[str]) -> List[str]:
        h = header[:]
        i = 0
        while i < len(h) - 1:
            cur = clean_str(h[i]); nxt = clean_str(h[i + 1])
            kcur = keyize(cur);    knxt = keyize(nxt)
            if ("ce" in kcur and ("cdfdpmo" in kcur or "cdf" in kcur)) and (knxt == "" or nxt == ""):
                h[i] = "CE"
                h[i + 1] = "CDF DPMO"
                i += 2
                continue
            i += 1
        return h

    def build_df_with_header(header: List[str], raw_rows: List[List[Any]]) -> pd.DataFrame:
        header = [norm_col(h) for h in header]
        header = split_combo_headers(header)
        hlen = len(header)
        fixed = []
        for r in raw_rows:
            rr = list(r or [])
            if len(rr) < hlen:
                rr = rr + [""] * (hlen - len(rr))
            elif len(rr) > hlen:
                rr = rr[:hlen]
            fixed.append(rr)
        df = pd.DataFrame(fixed, columns=header)

        rename_map = {}
        for col in df.columns:
            kcol = keyize(col)
            for ksrc, vdst in HEADER_MAP.items():
                if keyize(ksrc) == kcol:
                    rename_map[col] = vdst
                    break
        if rename_map:
            df = df.rename(columns=rename_map)

        if get_col(df, "CDF DPMO") is None:
            df = merge_adjacent_cols_if_needed(df, "CDF DPMO")
        if get_col(df, "LoR DPMO") is None:
            df = merge_adjacent_cols_if_needed(df, "LoR DPMO")
        if get_col(df, "DNR DPMO") is None:
            df = merge_adjacent_cols_if_needed(df, "DNR DPMO")

        print("PDF table columns:", list(df.columns))
        return df

    def split_ce_cdf_cell(val):
        if val is None:
            return (None, None)
        s = clean_str(val)
        nums = re.findall(r"[0-9.,\-]+", s)
        if len(nums) >= 2:
            return (to_num(nums[0]), to_num(nums[1]))
        return (to_num(s), None)

    rows: List[Dict[str, Any]] = []
    last_header: List[str] | None = None

    for page in pdf.pages:
        tables = page.extract_tables() or []
        for table in tables:
            if not table or len(table) < 1:
                continue

            data_rows_override: List[List[Any]] | None = None

            raw0 = [clean_str(h) for h in table[0]]
            raw1 = [clean_str(h) for h in table[1]] if len(table) > 1 else None

            header: List[str] | None = None
            data_start = 1

            if "Transporter ID" in raw0 or is_header_like(raw0):
                if raw1 and is_header_like(raw1):
                    combined = []
                    for a, b in zip(raw0, raw1):
                        if a and b: combined.append(f"{a} {b}".strip())
                        elif b:     combined.append(b)
                        else:       combined.append(a)
                    header = combined
                    data_start = 2
                else:
                    header = raw0
                    data_start = 1
                last_header = header[:]
            elif last_header is not None and table and looks_like_transporter_id(table[0][0]):
                header = last_header[:]
                data_start = 0

            if header is None:
                # FALLBACK: table without vertical rulings -> one cell per row.
                # Only reached when the regular column detection found nothing,
                # so PDFs with real columns keep their existing behaviour.
                fb = _parse_columnless_driver_table(page, table)
                if fb is None:
                    continue
                header, data_rows_override = fb

            df = build_df_with_header(
                header,
                data_rows_override if data_rows_override is not None else table[data_start:],
            )

            tid_col = get_col(df, "Transporter ID")
            if tid_col is None:
                continue

            delivered_c = get_col(df, "Delivered")
            dcr_c       = get_col(df, "DCR")
            pod_c       = get_col(df, "POD")
            cc_c        = get_col(df, "CC")
            ce_c        = get_col(df, "CE")
            lor_c       = get_col(df, "LoR DPMO")
            dnr_c       = get_col(df, "DNR DPMO")

            cdf_pct_c   = get_col(df, "CDF")
            cdf_dpmo_c  = get_col(df, "CDF DPMO")
            cdf_mode = None
            if cdf_pct_c is not None:
                cdf_mode = "PCT"
            elif cdf_dpmo_c is not None:
                cdf_mode = "DPMO"

            for _, r in df.iterrows():
                tid = clean_str(r.get(tid_col))
                if not tid or tid.lower() == "none":
                    continue

                ce_val = to_num(r.get(ce_c)) if ce_c else None
                cdf_dpmo_val = to_num(r.get(cdf_dpmo_c)) if cdf_dpmo_c else None
                cdf_pct_val  = to_percent(r.get(cdf_pct_c)) if cdf_pct_c else None

                if ce_c and cdf_dpmo_c is None and ce_val is None:
                    ce_try, cdf_try = split_ce_cdf_cell(r.get(ce_c))
                    if ce_val is None:
                        ce_val = ce_try
                    if cdf_dpmo_val is None:
                        cdf_dpmo_val = cdf_try

                rec = {
                    "Transporter ID": tid,
                    "Delivered": to_num(r.get(delivered_c)) if delivered_c else None,
                    "DCR":       to_percent(r.get(dcr_c))    if dcr_c else None,
                    "POD":       to_percent(r.get(pod_c))    if pod_c else None,
                    "CC":        to_percent(r.get(cc_c))     if cc_c else None,
                    "CE":        ce_val,
                    "LoR DPMO":  to_num(r.get(lor_c))        if lor_c else None,
                    "DNR DPMO":  to_num(r.get(dnr_c))        if dnr_c else None,
                    "CDF DPMO":  cdf_dpmo_val,
                    "CDF":       cdf_pct_val,
                    "_cdf_mode": cdf_mode,
                }

                # ONLY CHANGE: pass year into compute_scores so FinalScore regime uses (year, week)
                rec.update(compute_scores(rec, week_number=week_number, year=year))
                rows.append(rec)

    return rows

# ===============================
# SUMMARY EXTRACTION
# ===============================
def extract_summary(pdf: pdfplumber.PDF) -> Dict[str, Any]:
    text = "\n".join(filter(None, (p.extract_text() for p in pdf.pages))) or ""
    res: Dict[str, Any] = {}

    def grab(rex: str, flags=re.IGNORECASE):
        return re.search(rex, text, flags)

    if m := grab(r"\bWeek\s+(\d{1,2})\s*-\s*(\d{4})\b"):
        res["weekNumber"] = int(m.group(1))
        res["year"] = int(m.group(2))
        res["weekText"] = f"Week {m.group(1)} - {m.group(2)}"

    if m := grab(r"Overall\s+Score:\s*([\d.,]+)(?:\s*[|/]\s*([A-Za-z+\- ]+))?"):
        res["overallScore"] = to_num(m.group(1))
        if m.group(2):
            res["overallStatus"] = clean_str(m.group(2)).upper().replace(" ", "_")

    if res.get("overallStatus") is None:
        if m := grab(r"\bOverall\s+(?:Standing|Performance)\s+(?:as|is|:)?\s*([A-Za-z+\- ]+)\b"):
            res["overallStatus"] = clean_str(m.group(1)).upper().replace(" ", "_")

    if m := grab(r"Rank\s+at\s+([A-Z0-9\-]+)\s*:\s*(\d+)\s*\(\s*([+-]?\s*\d+)\s*WoW", re.IGNORECASE):
        res["stationCode"] = clean_str(m.group(1))
        res["rankAtStation"] = int(m.group(2))
        res["rankDeltaWoW"] = int(to_num(m.group(3)) or 0)

    if m := grab(r"Next\s+Day\s+Capacity\s+Reliability\s+([\d.,]+)\s*%"):
        res["reliabilityNextDay"] = to_num(m.group(1))

    if m := grab(r"(?:Same\s+Day|Sub-?Same\s+Day)[^%]*?Capacity\s+Reliability\s+([\d.,]+)\s*%"):
        res["reliabilitySameDay"] = to_num(m.group(1))

    if res.get("reliabilityNextDay") is not None:
        res["reliabilityScore"] = res["reliabilityNextDay"]

    # Neues Scorecard-Layout (seit ~KW29/2026): nur noch EINE Zeile
    # "Capacity Reliability 100.00%|Fantastic" ohne Next/Same-Day-Präfix.
    # Fallback nur, wenn die alten Muster nichts geliefert haben.
    if res.get("reliabilityScore") is None:
        if m := grab(r"Capacity\s+Reliability\s+([\d.,]+)\s*%"):
            res["reliabilityScore"] = to_num(m.group(1))

    if m := grab(r"(?:Rank\s+(?:in\s+Station|at\s+[A-Z0-9\-]+))\D*(\d+)\D*(?:of|/|von)\D*(\d+)", re.IGNORECASE):
        res["rankAtStation"] = int(m.group(1))
        res["stationCount"] = int(m.group(2))

    res.update(_parse_scorecard_summary_sections(pdf))

    return res

# ===============================
# POD QUALITY EXTRACTION
# ===============================
def _is_pod_quality_pdf(pdf: pdfplumber.PDF, filename: str) -> bool:
    name = (filename or "").lower()
    if any(tag in name for tag in ("pod-quality", "pod_quality", "photo_on_delivery")):
        return True

    # Check text on first two pages for common POD report signatures.
    first_pages_text = " ".join(
        clean_str(p.extract_text() or "").lower() for p in pdf.pages[:2]
    )
    text_signatures = (
        "photo on delivery quality report",
        "pod quality report",
        "photo on delivery quality",
    )
    if any(sig in first_pages_text for sig in text_signatures):
        return True

    # Fallback: detect POD summary tables by header patterns.
    for page in pdf.pages[:2]:
        for table in page.extract_tables() or []:
            if not table or not table[0]:
                continue
            header_cells = [clean_str(c).lower() for c in table[0]]
            header_text = " ".join(header_cells)
            has_category = "category" in header_cells or "category" in header_text
            has_pod_totals = (
                "total opportunities" in header_text or
                "total rejects" in header_text
            )
            if has_category and has_pod_totals:
                return True

    return False

def _extract_pod_quality_summary(pdf: pdfplumber.PDF, filename: str) -> Dict[str, Any]:
    res: Dict[str, Any] = {}
    text = clean_str(pdf.pages[0].extract_text() or "")

    if m := re.search(r"\b([A-Z0-9]{3,6})\s*-\s*Week\s*(\d{1,2})\b", text):
        res["stationCode"] = m.group(1)
        res["weekNumber"] = int(m.group(2))
        res["weekText"] = f"Week {m.group(2)}"

    if "year" not in res and filename:
        if m := re.search(r"(20\d{2})", filename):
            res["year"] = int(m.group(1))

    pod_summary: Dict[str, Any] = {}
    pod_rejects: Dict[str, Any] = {}

    tables = pdf.pages[0].extract_tables() or []
    for table in tables:
        if not table or len(table) < 2:
            continue
        header = [clean_str(c).lower() for c in table[0]]
        if "category" in header and any("total opportunities" in h for h in header):
            for row in table[1:]:
                cat = clean_str(row[0]).lower()
                cnt = to_num(row[1])
                pct = to_percent(row[2])
                key = keyize(cat)
                if key == "success":
                    pod_summary["successCount"] = cnt
                    pod_summary["successPct"] = pct
                elif key == "bypass":
                    pod_summary["bypassCount"] = cnt
                    pod_summary["bypassPct"] = pct
                elif key == "rejects":
                    pod_summary["rejectsCount"] = cnt
                    pod_summary["rejectsPct"] = pct
                elif key == "opportunities":
                    pod_summary["opportunitiesCount"] = cnt
                    pod_summary["opportunitiesPct"] = pct
        elif "category" in header and any("total rejects" in h for h in header):
            for row in table[1:]:
                cat = clean_str(row[0]).lower()
                cnt = to_num(row[1])
                pct = to_percent(row[2])
                key = keyize(cat)
                if key == "blurryphoto":
                    pod_rejects["blurryPhotoCount"] = cnt
                    pod_rejects["blurryPhotoPct"] = pct
                elif key == "phototoodark":
                    pod_rejects["photoTooDarkCount"] = cnt
                    pod_rejects["photoTooDarkPct"] = pct
                elif key == "nopackagedetected":
                    pod_rejects["noPackageDetectedCount"] = cnt
                    pod_rejects["noPackageDetectedPct"] = pct
                elif key == "packageincar":
                    pod_rejects["packageInCarCount"] = cnt
                    pod_rejects["packageInCarPct"] = pct
                elif key == "packagetooclose":
                    pod_rejects["packageTooCloseCount"] = cnt
                    pod_rejects["packageTooClosePct"] = pct
                elif key in ("grandtotal", "total"):
                    pod_rejects["totalRejectsCount"] = cnt
                    pod_rejects["totalRejectsPct"] = pct

    if pod_summary:
        res["podQualitySummary"] = pod_summary
    if pod_rejects:
        res["podQualityRejects"] = pod_rejects

    return res

def _extract_pod_quality_drivers(pdf: pdfplumber.PDF) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    def looks_like_transporter_id(x: Any) -> bool:
        s = clean_str(x)
        return bool(re.match(r"^[A-Z0-9]{8,20}$", s))

    for page in pdf.pages:
        tables = page.extract_tables() or []
        for table in tables:
            if not table:
                continue
            for row in table:
                tid = clean_str(row[0])
                if not looks_like_transporter_id(tid):
                    continue
                r = list(row) + [""] * 10
                rows.append(
                    {
                        "Transporter ID": tid,
                        "POD_Q_Opportunities": to_num(r[1]),
                        "POD_Q_Success": to_num(r[2]),
                        "POD_Q_Bypass": to_num(r[3]),
                        "POD_Q_Rejects": to_num(r[4]),
                        "POD_Q_BlurryPhoto": to_num(r[5]),
                        "POD_Q_NoPackageDetected": to_num(r[6]),
                        "POD_Q_PackageInCar": to_num(r[7]),
                        "POD_Q_PackageTooClose": to_num(r[8]),
                        "POD_Q_PhotoTooDark": to_num(r[9]),
                    }
                )

    return rows

# ===============================
# RANKING
# ===============================
def add_ranking_and_status(
    drivers: List[Dict[str, Any]],
    week_number: Optional[int] = None,
    year: Optional[int] = None,
) -> None:
    sortable = [d for d in drivers if isinstance(d.get("FinalScore"), (int, float))]
    sortable.sort(key=lambda x: x["FinalScore"], reverse=True)

    rank = 0
    last_score = None
    score_to_rank: Dict[float, int] = {}
    for d in sortable:
        fs = d["FinalScore"]
        if fs != last_score:
            rank += 1
            last_score = fs
        score_to_rank[fs] = rank

    for d in drivers:
        fs = d.get("FinalScore")
        d["rank"] = score_to_rank.get(fs) if isinstance(fs, (int, float)) else None
        d["statusBucket"] = status_bucket(fs, week_number=week_number, year=year)

# ===============================
# ROUTES
# ===============================
# ===============================
# FLEET (Cortex VehiclesData) XLSX HANDLING
# ===============================
#
# Amazon's Cortex portal exports the DSP fleet as an XLSX with German
# headers: FIN, Fahrzeugname, Nummernschild, Marke, Modell, Untermodell,
# Status, Betriebsstatus, Fahrzeuganbieter, Eigentumstypus,
# Startdatum/Enddatum des Eigentums, payload, cubicCapacity, ...
#
# We extract every row, normalize the key fields, and return a flat
# list. The Flutter side handles dedup by FIN/VIN.

# Cortex exportiert je nach Spracheinstellung des Kontos deutsche ODER
# englische Spaltenüberschriften. Beide Fassungen müssen erkannt werden —
# vorher scheiterte der englische Export mit "Keine Fahrzeuge erkannt".
VEHICLES_REQUIRED_HEADERS = {"fin", "nummernschild"}
VEHICLES_REQUIRED_HEADERS_EN = {"vin", "licenseplatenumber"}

# Zielfeld → mögliche Spaltenüberschriften (deutsch zuerst, dann englisch).
VEHICLE_COLUMN_ALIASES = {
    "vinNumber": ["FIN", "vin"],
    "plateNumber": ["Nummernschild", "licensePlateNumber"],
    "fleetName": ["Fahrzeugname", "vehicleName"],
    "brand": ["Marke", "make"],
    "model": ["Modell", "model"],
    "submodel": ["Untermodell", "subModel"],
    "status": ["Status"],
    "statusPriority": ["Status-Priorität", "statusPriority"],
    "statusReasonCode": ["Status-Ursachencode", "statusReasonCode"],
    "statusReasonMessage": ["StatusGrundMeldung", "statusReasonMessage"],
    "operationalStatus": ["Betriebsstatus", "operationalStatus"],
    "statusSearchValue": ["StatusSuchwert", "statusSearchValue"],
    "subcontractorName": ["subcontractorName"],
    "vehicleProvider": ["Fahrzeuganbieter", "vehicleProvider"],
    "registrationType": [
        "Fahrzeugregistrierungstypus", "vehicleRegistrationType",
    ],
    "manufacturingYear": ["Jahr", "year"],
    "vehicleType": ["Typus", "Type"],
    "ownershipType": ["Eigentumstypus", "ownershipType"],
    "ownershipStart": ["Startdatum des Eigentums", "ownershipStartDate"],
    "ownershipEnd": ["Enddatum des Eigentums", "ownershipEndDate"],
    "pmStatistics": ["PM-Statistiken", "pmStats"],
    "registrationExpiry": [
        "Datum des Ablaufs der Registrierung", "registrationExpiryDate",
    ],
    "registeredState": ["registeredState"],
    "serviceTier": ["serviceTier"],
    "stationCode": ["stationCode"],
    "payload": ["payload"],
    "cubicCapacity": ["cubicCapacity"],
    "serviceType": ["serviceType"],
}


def _is_vehicles_xlsx(filename: str, content: bytes) -> bool:
    name = (filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls")):
        return False
    # Filename hint
    if "vehicles" in name or "fahrzeug" in name or "cortex" in name:
        return True
    # Header sniff
    try:
        from openpyxl import load_workbook as _peek
        wb = _peek(io.BytesIO(content), data_only=True, read_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            first = next(ws.iter_rows(values_only=True), None)
            if not first:
                continue
            heads = {str(c or "").strip().lower() for c in first}
            if (VEHICLES_REQUIRED_HEADERS.issubset(heads)
                    or VEHICLES_REQUIRED_HEADERS_EN.issubset(heads)):
                return True
    except Exception:
        pass
    return False


def _extract_vehicles_from_xlsx(content: bytes, filename: str):
    """Returns ({}, [vehicles…]).
    Each vehicle is a flat camelCase dict ready for Firestore."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    # Build header → index map (German lowercased)
    headers = [str(c or "").strip() for c in next(ws.iter_rows(values_only=True))]
    idx = {h.lower(): i for i, h in enumerate(headers)}

    def cell(row, name: str) -> str:
        """Liest eine Spalte. `name` ist entweder eine Überschrift oder ein
        Zielfeld aus VEHICLE_COLUMN_ALIASES; im zweiten Fall werden die
        deutsche und die englische Überschrift der Reihe nach probiert."""
        candidates = VEHICLE_COLUMN_ALIASES.get(name, [name])
        for cand in candidates:
            i = idx.get(cand.lower())
            if i is None or i >= len(row):
                continue
            v = row[i]
            if v is None:
                continue
            text = str(v).strip()
            if text:
                return text
        return ""

    def cell_int(row, name: str):
        v = cell(row, name)
        if not v:
            return None
        try:
            return int(float(v))
        except ValueError:
            return None

    vehicles: List[Dict[str, Any]] = []
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri == 1:
            continue  # skip header
        fin = cell(row, "vinNumber")
        plate = cell(row, "plateNumber")
        if not fin and not plate:
            continue  # empty row

        vehicles.append({
            "vinNumber": fin,
            "plateNumber": plate,
            "fleetName": cell(row, "fleetName"),
            "brand": cell(row, "brand"),
            "model": cell(row, "model"),
            "submodel": cell(row, "submodel"),
            "status": cell(row, "status"),
            "statusPriority": cell(row, "statusPriority"),
            "statusReasonCode": cell(row, "statusReasonCode"),
            "statusReasonMessage": cell(row, "statusReasonMessage"),
            "operationalStatus": cell(row, "operationalStatus"),
            "statusSearchValue": cell(row, "statusSearchValue"),
            "subcontractorName": cell(row, "subcontractorName"),
            "vehicleProvider": cell(row, "vehicleProvider"),
            "registrationType": cell(row, "registrationType"),
            "manufacturingYear": cell_int(row, "manufacturingYear"),
            "vehicleType": cell(row, "vehicleType"),
            "ownershipType": cell(row, "ownershipType"),
            "ownershipStart": cell(row, "ownershipStart"),
            "ownershipEnd": cell(row, "ownershipEnd"),
            "pmStatistics": cell(row, "pmStatistics"),
            "registrationExpiry": cell(row, "registrationExpiry"),
            "registeredState": cell(row, "registeredState"),
            "serviceTier": cell(row, "serviceTier"),
            "stationCode": cell(row, "stationCode"),
            "payload": cell(row, "payload"),
            "cubicCapacity": cell(row, "cubicCapacity"),
            "serviceType": cell(row, "serviceType"),
            "_source": "cortex_xlsx",
        })

    summary = {
        "vehiclesImport": {
            "totalRows": len(vehicles),
            "source": "cortex",
        }
    }
    return summary, vehicles


# ===============================
# CONCESSIONS XLSX HANDLING
# ===============================

def _is_concessions_xlsx(filename: str) -> bool:
    name = (filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls")):
        return False
    return "concession" in name


_FOCUS_BUCKET_MAP = {
    "attended dnr deliveries": "attendedDnr",
    "photo on delivery": "photoOnDelivery",
    "successful contact opportunity": "successfulContact",
    "delivered > 25m": "delivered25m",
    "feedback false scan indicator": "falseScan",
    "mailbox eligible, delivered elsewhere": "mailbox",
    "delivered using otp": "deliveredOtp",
}

_FOCUS_BUCKET_TX_KEYS = {
    "attended dnr deliveries": "CONC_Focus_AttendedDnr",
    "photo on delivery": "CONC_Focus_PhotoOnDelivery",
    "successful contact opportunity": "CONC_Focus_SuccessfulContact",
    "delivered > 25m": "CONC_Focus_Delivered25m",
    "feedback false scan indicator": "CONC_Focus_FalseScan",
    "mailbox eligible, delivered elsewhere": "CONC_Focus_Mailbox",
    "delivered using otp": "CONC_Focus_DeliveredOtp",
}


def _xlsx_num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if not s or s == "—":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_concessions_meta_from_filename(filename: str) -> Dict[str, Any]:
    """Pull country/dsp/station/week from a filename like
    'DE-AION-DBY5-Week19-Concessions.xlsx'."""
    out: Dict[str, Any] = {}
    if not filename:
        return out
    base = filename.rsplit("/", 1)[-1]
    base = base.rsplit(".", 1)[0]
    m = re.search(r"([A-Z]{2})-([A-Z0-9]+)-([A-Z0-9]+)-Week(\d+)", base, re.I)
    if m:
        out["country"] = m.group(1).upper()
        out["dsp"] = m.group(2).upper()
        out["stationCode"] = m.group(3).upper()
        out["weekNumber"] = int(m.group(4))
    return out


def _extract_concessions_from_xlsx(content: bytes, filename: str):
    """Returns (summary_dict, drivers_list)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheets = {s.lower(): s for s in wb.sheetnames}

    meta = _parse_concessions_meta_from_filename(filename)

    # ----- DSP-level: Focus Areas DSP block -----
    focus_dsp: Dict[str, float] = {}
    week_labels: List[str] = []
    if "focus areas" in sheets:
        ws = wb[sheets["focus areas"]]
        rows = list(ws.iter_rows(values_only=True))
        in_dsp_block = False
        for row in rows:
            if not row:
                continue
            first = row[0]
            if first is None:
                continue
            text = str(first).strip()
            if text == "Bucket":
                in_dsp_block = True
                week_labels = [
                    str(c).strip() for c in row[1:5]
                    if c is not None
                ]
                continue
            if not in_dsp_block:
                continue
            if text.lower() == "grand total":
                break
            key = _FOCUS_BUCKET_MAP.get(text.lower())
            if key:
                total = _xlsx_num(row[5]) if len(row) > 5 else None
                if total is not None:
                    focus_dsp[key] = total

    # ----- DSP-level: DNR by DSP -----
    del_volume = dnr_count = dnr_dpmo = None
    week_number = meta.get("weekNumber")
    year = None
    if "dnr by dsp" in sheets:
        ws = wb[sheets["dnr by dsp"]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows: List[tuple] = []
        for row in rows:
            if not row or row[0] in (None, "DNR by DSP", "dsp"):
                continue
            data_rows.append(row)
        if data_rows:
            last = data_rows[-1]
            del_volume = _xlsx_num(last[2]) if len(last) > 2 else None
            dnr_count = _xlsx_num(last[3]) if len(last) > 3 else None
            dnr_dpmo = _xlsx_num(last[4]) if len(last) > 4 else None
            # week label like '2026-19'
            wk = str(last[1] or "").strip()
            m = re.match(r"(\d{4})-(\d{1,2})", wk)
            if m:
                year = int(m.group(1))
                if week_number is None:
                    week_number = int(m.group(2))

    summary_block = {
        "delVolume": del_volume,
        "dnrCount": dnr_count,
        "dnrDpmo": dnr_dpmo,
        "weekLabels": week_labels,
    }

    # ----- Per driver: DNR by Transporter ID -----
    drivers_by_id: Dict[str, Dict[str, Any]] = {}
    if "dnr by transporter id" in sheets:
        ws = wb[sheets["dnr by transporter id"]]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[0] == "Transporter ID":
                header_idx = i
                break
        if header_idx is not None:
            for row in rows[header_idx + 1:]:
                if not row or not row[0]:
                    continue
                tid = str(row[0]).strip()
                if tid.lower() in ("grand total", "total"):
                    continue
                drivers_by_id[tid] = {
                    "Transporter ID": tid,
                    "CONC_TotalDelivered": _xlsx_num(row[1]) if len(row) > 1 else None,
                    "CONC_DnrCount_W1": _xlsx_num(row[2]) if len(row) > 2 else None,
                    "CONC_DnrCount_W2": _xlsx_num(row[3]) if len(row) > 3 else None,
                    "CONC_DnrCount_W3": _xlsx_num(row[4]) if len(row) > 4 else None,
                    "CONC_DnrCount_W4": _xlsx_num(row[5]) if len(row) > 5 else None,
                    "CONC_TotalDnr": _xlsx_num(row[6]) if len(row) > 6 else None,
                    "CONC_DnrDpmo_W1": _xlsx_num(row[7]) if len(row) > 7 else None,
                    "CONC_DnrDpmo_W2": _xlsx_num(row[8]) if len(row) > 8 else None,
                    "CONC_DnrDpmo_W3": _xlsx_num(row[9]) if len(row) > 9 else None,
                    "CONC_DnrDpmo_W4": _xlsx_num(row[10]) if len(row) > 10 else None,
                    "CONC_DnrDpmo4w": _xlsx_num(row[11]) if len(row) > 11 else None,
                }

    # ----- Per driver: Focus Areas Transporter block (totals) -----
    if "focus areas" in sheets:
        ws = wb[sheets["focus areas"]]
        rows = list(ws.iter_rows(values_only=True))
        in_tx_block = False
        for row in rows:
            if not row:
                continue
            first = row[0]
            if first is None:
                continue
            text = str(first).strip()
            if text == "Transporter ID":
                in_tx_block = True
                continue
            if not in_tx_block:
                continue
            if text.lower() == "grand total":
                continue
            # rows: tid | bucket | w1 | w2 | w3 | w4 | total
            tid = text
            bucket_label = str(row[1] or "").strip().lower() if len(row) > 1 else ""
            if not bucket_label or bucket_label == "total":
                continue
            key = _FOCUS_BUCKET_TX_KEYS.get(bucket_label)
            if not key:
                continue
            total = _xlsx_num(row[6]) if len(row) > 6 else None
            if total is None:
                # Some rows place 4-week sum at column 5 instead of 6;
                # fall back to sum of weekly cols.
                week_sum = sum(
                    (_xlsx_num(row[i]) or 0) for i in range(2, min(6, len(row)))
                )
                total = week_sum if week_sum else None
            if total is None:
                continue
            entry = drivers_by_id.setdefault(tid, {"Transporter ID": tid})
            entry[key] = (entry.get(key) or 0) + total

    drivers = list(drivers_by_id.values())

    summary: Dict[str, Any] = {
        "concessionsSummary": summary_block,
        "concessionsFocusBuckets": focus_dsp,
        "weekText": f"KW {week_number}" if week_number else None,
        "weekNumber": week_number,
        "year": year,
        "stationCode": meta.get("stationCode"),
    }
    # strip None values from summary for a tidy response
    summary = {k: v for k, v in summary.items() if v is not None}

    return summary, drivers


# ===============================
# NEW DSC CONCESSION XLSX HANDLING (Amazon 2026 format)
# ===============================

# Maps the new DSC bucket labels (case-insensitive, partial OK) to internal keys.
# These replace the older "focus areas" buckets.
_DSC_BUCKET_MAP = {
    "delivered to household member / customer": "deliveredToHouseholdMember",
    "delivered to household member/customer": "deliveredToHouseholdMember",
    "delivered to household member": "deliveredToHouseholdMember",
    "delivery preferences not followed": "deliveryPreferencesNotFollowed",
    "multiple concessions reasons": "multipleConcessionsReasons",
    "geo distance > 25m": "geoDistance25m",
    "geo distance > 25 m": "geoDistance25m",
    "delivered to neighbour": "deliveredToNeighbour",
    "delivered to neighbor": "deliveredToNeighbour",
    "delivered to receptionist": "deliveredToReceptionist",
}


def _is_dsc_concessions_xlsx_by_sheets(wb) -> bool:
    """Detect new DSC format by looking for DSC-specific sheet names."""
    sheets_lower = {s.lower() for s in wb.sheetnames}
    return (
        "dsc concessions" in sheets_lower
        or "dsc by transporter id" in sheets_lower
        or "dsc concession bucket summary" in sheets_lower
    )


def _fmt_xlsx_date(v: Any) -> Optional[str]:
    """Convert Excel cell value to ISO date string. Returns None if not a
    valid date/datetime."""
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            pass
    s = str(v).strip()
    if not s or s == "—":
        return None
    return s


def _extract_dsc_concessions_from_xlsx(content: bytes, filename: str):
    """Parse Amazon DSC Concession XLSX (new 2026 format).

    Returns (summary_dict, drivers_list) — schema-compatible with the
    legacy `_extract_concessions_from_xlsx` so existing UI keeps working,
    plus additional fields:
      * CONC_DnrEvents — per-driver list of {trackingId, dnrDate,
                          deliveryDateTime, ...} for exact-date display
      * CONC_DnrCountByWeek — actual {year-week: count} map
      * CONC_DscBuckets — {bucketKey: count} per driver
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheets = {s.lower(): s for s in wb.sheetnames}
    meta = _parse_concessions_meta_from_filename(filename)

    drivers_by_id: Dict[str, Dict[str, Any]] = {}
    week_labels: List[str] = []

    # ----- Sheet 1: "DSC by Transporter ID" (per-driver weekly counts + DPMO) -----
    if "dsc by transporter id" in sheets:
        ws = wb[sheets["dsc by transporter id"]]
        rows = list(ws.iter_rows(values_only=True))

        header_idx = None
        for i, row in enumerate(rows):
            if row and any(
                c is not None and str(c).strip() == "Transporter ID"
                for c in row
            ):
                header_idx = i
                break

        if header_idx is not None:
            headers = rows[header_idx]
            count_cols: List[tuple] = []   # (col_idx, "YYYY-WW")
            dpmo_cols: List[tuple] = []
            total_delivered_col: Optional[int] = None
            total_dsc_count_col: Optional[int] = None
            dpmo_4w_col: Optional[int] = None

            for j, h in enumerate(headers):
                if h is None:
                    continue
                hs = str(h).strip()
                m_count = re.match(r"^(\d{4})-(\d{1,2})_DSC_Count$", hs, re.I)
                m_dpmo = re.match(r"^(\d{4})-(\d{1,2})_DSC_DPMO$", hs, re.I)
                if m_count:
                    count_cols.append(
                        (j, f"{m_count.group(1)}-{m_count.group(2).zfill(2)}")
                    )
                elif m_dpmo:
                    dpmo_cols.append(
                        (j, f"{m_dpmo.group(1)}-{m_dpmo.group(2).zfill(2)}")
                    )
                elif hs.lower() == "total delivered":
                    total_delivered_col = j
                elif hs.lower() == "total dsc count":
                    total_dsc_count_col = j
                elif "last 4 weeks" in hs.lower() or "last4weeks" in hs.lower():
                    dpmo_4w_col = j

            count_cols.sort(key=lambda x: x[1])
            dpmo_cols.sort(key=lambda x: x[1])
            week_labels = [w for _, w in count_cols]

            for row in rows[header_idx + 1:]:
                if not row or row[0] is None:
                    continue
                tid = str(row[0]).strip()
                if not tid:
                    continue
                tid_lower = tid.lower()
                # Filter out totals/summary rows (e.g. "Total DSC Count",
                # "Total Delivered", "Grand Total", "Total")
                if (
                    tid_lower in ("grand total", "total")
                    or tid_lower.startswith("total ")
                    or "total dsc" in tid_lower
                    or "grand total" in tid_lower
                ):
                    continue

                entry: Dict[str, Any] = {"Transporter ID": tid}

                if total_delivered_col is not None and len(row) > total_delivered_col:
                    entry["CONC_TotalDelivered"] = _xlsx_num(
                        row[total_delivered_col]
                    )
                if total_dsc_count_col is not None and len(row) > total_dsc_count_col:
                    entry["CONC_TotalDnr"] = _xlsx_num(row[total_dsc_count_col])
                if dpmo_4w_col is not None and len(row) > dpmo_4w_col:
                    entry["CONC_DnrDpmo4w"] = _xlsx_num(row[dpmo_4w_col])

                # Backwards-compatible W1..W4 (oldest..newest of available weeks)
                latest4_counts = count_cols[-4:]
                latest4_dpmos = dpmo_cols[-4:]
                for idx, (col_i, _) in enumerate(latest4_counts):
                    entry[f"CONC_DnrCount_W{idx + 1}"] = (
                        _xlsx_num(row[col_i]) if len(row) > col_i else None
                    )
                for idx, (col_i, _) in enumerate(latest4_dpmos):
                    entry[f"CONC_DnrDpmo_W{idx + 1}"] = (
                        _xlsx_num(row[col_i]) if len(row) > col_i else None
                    )

                # Exact weekly counts keyed by real year-week label
                by_week: Dict[str, float] = {}
                for col_i, week_lbl in count_cols:
                    val = _xlsx_num(row[col_i]) if len(row) > col_i else None
                    if val is not None:
                        by_week[week_lbl] = val
                if by_week:
                    entry["CONC_DnrCountByWeek"] = by_week

                drivers_by_id[tid] = entry

    # ----- Sheet 2: "DSC Concessions" (event-level rows with dates) -----
    events_by_driver: Dict[str, List[Dict[str, Any]]] = {}
    if "dsc concessions" in sheets:
        ws = wb[sheets["dsc concessions"]]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers_lower = [
                str(c or "").strip().lower() for c in rows[0]
            ]

            def hcol(name: str) -> Optional[int]:
                try:
                    return headers_lower.index(name)
                except ValueError:
                    return None

            c_tid = hcol("transporter_id")
            c_tracking = hcol("tracking_id")
            c_dnr_date = hcol("dnr_date")
            c_delivery_dt = hcol("delivery_date_time")
            c_actual = hcol("actual_delivery_type")
            c_promised = hcol("promised_delivery_type")
            c_year = hcol("year")
            c_week = hcol("week")

            for row in rows[1:]:
                if not row or c_tid is None or len(row) <= c_tid or row[c_tid] is None:
                    continue
                tid = str(row[c_tid]).strip()
                if not tid:
                    continue

                def at(idx):
                    return row[idx] if idx is not None and len(row) > idx else None

                evt = {
                    "trackingId": (str(at(c_tracking)).strip()
                                   if at(c_tracking) is not None else None),
                    "dnrDate": _fmt_xlsx_date(at(c_dnr_date)),
                    "deliveryDateTime": _fmt_xlsx_date(at(c_delivery_dt)),
                    "actualDeliveryType": (str(at(c_actual)).strip()
                                            if at(c_actual) is not None else None),
                    "promisedDeliveryType": (str(at(c_promised)).strip()
                                              if at(c_promised) is not None else None),
                    "year": (int(at(c_year))
                              if at(c_year) is not None
                              and str(at(c_year)).strip().isdigit() else None),
                    "week": (int(at(c_week))
                              if at(c_week) is not None
                              and str(at(c_week)).strip().isdigit() else None),
                }
                # strip None
                evt = {k: v for k, v in evt.items() if v is not None}
                events_by_driver.setdefault(tid, []).append(evt)

    for tid, events in events_by_driver.items():
        entry = drivers_by_id.setdefault(tid, {"Transporter ID": tid})
        # Sort newest first
        def evt_sort_key(e):
            return e.get("dnrDate") or e.get("deliveryDateTime") or ""
        events.sort(key=evt_sort_key, reverse=True)
        entry["CONC_DnrEvents"] = events

    # ----- Sheet 3: "DSC Concession Bucket Summary" (DSP + per-driver buckets) -----
    focus_dsp: Dict[str, float] = {}
    buckets_by_driver: Dict[str, Dict[str, float]] = {}

    bucket_sheet_name = None
    for s_lower, s_orig in sheets.items():
        if "dsc" in s_lower and "bucket" in s_lower and "summary" in s_lower:
            bucket_sheet_name = s_orig
            break

    if bucket_sheet_name:
        ws = wb[bucket_sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        in_dsp_block = False
        in_tid_block = False
        grand_total_col: Optional[int] = None

        for row in rows:
            if not row or row[0] is None:
                continue
            first = str(row[0]).strip()
            if not first:
                continue
            first_lower = first.lower()

            if first_lower == "bucket":
                in_dsp_block = True
                in_tid_block = False
                grand_total_col = None
                for j, h in enumerate(row):
                    if h is not None and str(h).strip().lower() == "grand total":
                        grand_total_col = j
                        break
                continue

            if first_lower == "transporter id":
                in_dsp_block = False
                in_tid_block = True
                grand_total_col = None
                for j, h in enumerate(row):
                    if h is not None and str(h).strip().lower() == "grand total":
                        grand_total_col = j
                        break
                continue

            if first_lower in ("grand total", "total"):
                continue

            if in_dsp_block:
                bucket_key = _DSC_BUCKET_MAP.get(first_lower)
                if bucket_key and grand_total_col is not None and len(row) > grand_total_col:
                    val = _xlsx_num(row[grand_total_col])
                    if val is not None:
                        focus_dsp[bucket_key] = (
                            focus_dsp.get(bucket_key, 0) + val
                        )

            elif in_tid_block:
                # row: TID | Bucket | year-week cols... | Grand Total
                tid = first
                bucket_label_raw = (
                    str(row[1] or "").strip().lower() if len(row) > 1 else ""
                )
                bucket_key = _DSC_BUCKET_MAP.get(bucket_label_raw)
                if bucket_key and grand_total_col is not None and len(row) > grand_total_col:
                    val = _xlsx_num(row[grand_total_col])
                    if val is not None:
                        d_buckets = buckets_by_driver.setdefault(tid, {})
                        d_buckets[bucket_key] = d_buckets.get(bucket_key, 0) + val

    for tid, buckets in buckets_by_driver.items():
        entry = drivers_by_id.setdefault(tid, {"Transporter ID": tid})
        entry["CONC_DscBuckets"] = buckets
        for key, val in buckets.items():
            entry[f"CONC_Bucket_{key}"] = val

    # ----- DSP-level summary aggregates -----
    total_dsc_count = 0.0
    total_delivered_dsp = 0.0
    for d in drivers_by_id.values():
        td = d.get("CONC_TotalDelivered")
        if isinstance(td, (int, float)):
            total_delivered_dsp += td
        tc = d.get("CONC_TotalDnr")
        if isinstance(tc, (int, float)):
            total_dsc_count += tc

    dnr_dpmo = (
        (total_dsc_count / total_delivered_dsp * 1_000_000)
        if total_delivered_dsp > 0
        else None
    )

    year: Optional[int] = None
    week_number = meta.get("weekNumber")
    if week_labels:
        latest = week_labels[-1]
        m = re.match(r"(\d{4})-(\d{1,2})", latest)
        if m:
            year = int(m.group(1))
            if week_number is None:
                week_number = int(m.group(2))

    summary_block = {
        "delVolume": total_delivered_dsp if total_delivered_dsp > 0 else None,
        "dnrCount": total_dsc_count if total_dsc_count > 0 else None,
        "dnrDpmo": dnr_dpmo,
        "weekLabels": week_labels,
    }
    summary_block = {k: v for k, v in summary_block.items() if v is not None}

    drivers = list(drivers_by_id.values())

    summary: Dict[str, Any] = {
        "concessionsSummary": summary_block,
        "concessionsFocusBuckets": focus_dsp,
        "weekText": f"KW {week_number}" if week_number else None,
        "weekNumber": week_number,
        "year": year,
        "stationCode": meta.get("stationCode"),
        "concessionsFormat": "dsc",
    }
    summary = {k: v for k, v in summary.items() if v is not None}

    return summary, drivers


# =================================================================
#  CDF (Customer Delivery Feedback) HTML report parsing
# =================================================================
#
# Amazon's CDF report comes as an HTML file with a single big <tbody>
# of "Negative Feedback Tracking IDs". Each row has the columns:
#
#   Tracking ID | Transporter ID | Feedback L0 | L1 | L2 |
#   Feedback Date | Delivery Time | City | Postal Code |
#   Contact Compliance | PHR Compliance | PHR Safe Place |
#   PHR Delivery Location | Package Scanned > 25m Distance |
#   DNR Concession
#
# We aggregate per Transporter ID and keep all event-level rows so the
# UI can show "which feedbacks happened on which dates" per driver.

# Stable mapping from raw "Feedback L1" labels to camelCase keys used
# in Firestore + the Flutter UI. Anything unknown falls into "other".
CDF_L1_KEY_MAP: Dict[str, str] = {
    "never received delivery": "neverReceived",
    "driver mishandled package": "driverMishandled",
    "not delivered to preferred location": "notDeliveredToPreferredLocation",
    "damaged package": "damagedPackage",
    "delivery was late": "deliveryWasLate",
    "bad delivery experience": "badDeliveryExperience",
    "driver was unprofessional": "driverWasUnprofessional",
}


def _is_cdf_html(filename: str, content: bytes) -> bool:
    name = (filename or "").lower()
    if "cdf" in name and (
        name.endswith(".html") or name.endswith(".htm")
    ):
        return True
    # Sniff first few KB of the file — Amazon's title includes "CDF Report"
    head = content[:4096].decode("utf-8", errors="ignore").lower()
    return ("cdf report" in head) and ("<html" in head or "<table" in head)


def _parse_cdf_meta_from_filename(filename: str) -> Dict[str, Any]:
    """e.g. "DE-AION-DBY5-Week19-CDF-report.html"
    → {country: 'DE', dsp: 'AION', stationCode: 'DBY5',
       weekNumber: 19, year: <inferred>}.
    """
    out: Dict[str, Any] = {}
    if not filename:
        return out
    base = filename.rsplit("/", 1)[-1]
    m = re.search(r"([A-Z]{2})-([A-Z]+)-([A-Z0-9]+)-Week(\d{1,2})", base, re.I)
    if m:
        out["country"] = m.group(1).upper()
        out["dsp"] = m.group(2).upper()
        out["stationCode"] = m.group(3).upper()
        out["weekNumber"] = int(m.group(4))
    y = re.search(r"(20\d{2})", base)
    if y:
        out["year"] = int(y.group(1))
    return out


def _extract_cdf_from_html(content: bytes, filename: str):
    """Parse the negative-feedback table → per-driver aggregations
    plus the raw event rows so the UI can render per-event details.
    """
    try:
        from bs4 import BeautifulSoup  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "CDF HTML parsing requires beautifulsoup4 — add it to "
            "requirements.txt and redeploy."
        ) from e

    meta = _parse_cdf_meta_from_filename(filename)

    html = content.decode("utf-8", errors="replace")
    soup = BeautifulSoup(html, "html.parser")

    # Find the table with the "Tracking ID" header — there is only one
    # data table of interest in the report. We also identify the actual
    # header row (some Amazon variants stack a thead + tfoot which would
    # duplicate th cells if naively flattened).
    target_table = None
    header_cells: List[str] = []
    for table in soup.find_all("table"):
        # Look at each header row's *direct* th children only (avoid
        # nested tables and tfoot duplicates).
        for tr in table.find_all("tr"):
            ths = tr.find_all("th", recursive=False)
            labels = [(th.get_text() or "").strip() for th in ths]
            lower = [s.lower() for s in labels]
            if "tracking id" in lower and "transporter id" in lower:
                target_table = table
                header_cells = labels
                break
        if target_table is not None:
            break
    if target_table is None or not header_cells:
        # No data table → return empty result, but still pass meta back so
        # the UI can record "uploaded empty week" if needed.
        return meta, []

    # Build a column-name → index map (lowercased keys)
    col_idx: Dict[str, int] = {
        h.lower(): i for i, h in enumerate(header_cells)
    }

    def cell_text(row_cells, name: str) -> str:
        i = col_idx.get(name.lower())
        if i is None or i >= len(row_cells):
            return ""
        return (row_cells[i].get_text() or "").strip()

    # Aggregate by transporter id
    by_driver: Dict[str, Dict[str, Any]] = {}
    events: List[Dict[str, Any]] = []

    for tr in target_table.find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue
        tid = cell_text(tds, "Transporter ID")
        if not tid:
            continue
        l0 = cell_text(tds, "Feedback L0")
        l1 = cell_text(tds, "Feedback L1")
        l2 = cell_text(tds, "Feedback L2")
        fb_date = cell_text(tds, "Feedback Date")
        delivery_time = cell_text(tds, "Delivery Time")
        city = cell_text(tds, "City")
        postal = cell_text(tds, "Postal Code")
        tracking = cell_text(tds, "Tracking ID")
        dnr_concession = cell_text(tds, "DNR Concession")
        phr_loc = cell_text(tds, "PHR - Delivery Location")
        far_scan = cell_text(tds, "Package Scanned > 25m Distance")

        event = {
            "trackingId": tracking,
            "transporterId": tid,
            "feedbackL0": l0,
            "feedbackL1": l1,
            "feedbackL2": l2,
            "feedbackDate": fb_date,
            "deliveryTime": delivery_time,
            "city": city,
            "postalCode": postal,
            "phrDeliveryLocation": phr_loc,
            "packageScannedFar": (far_scan or "").upper() == "Y",
            "dnrConcession": (dnr_concession or "").upper() == "Y",
        }
        events.append(event)

        d = by_driver.setdefault(tid, {
            "Transporter ID": tid,
            "CDF_TotalFeedback": 0,
            "CDF_NegativeFeedback": 0,
            "CDF_L1_NeverReceived": 0,
            "CDF_L1_DriverMishandled": 0,
            "CDF_L1_NotDeliveredToPreferredLocation": 0,
            "CDF_L1_DamagedPackage": 0,
            "CDF_L1_DeliveryWasLate": 0,
            "CDF_L1_BadDeliveryExperience": 0,
            "CDF_L1_DriverWasUnprofessional": 0,
            "CDF_L1_Other": 0,
            "CDF_Events": [],
        })
        d["CDF_TotalFeedback"] += 1
        if (l0 or "").lower().startswith("was not so great"):
            d["CDF_NegativeFeedback"] += 1
        # Bucketize by L1
        key = CDF_L1_KEY_MAP.get((l1 or "").lower(), None)
        if key:
            pascal = key[0].upper() + key[1:]
            field = f"CDF_L1_{pascal}"
            if field in d:
                d[field] += 1
            else:
                d["CDF_L1_Other"] += 1
        else:
            d["CDF_L1_Other"] += 1
        d["CDF_Events"].append(event)

    # Distribution across all drivers (for the hero card)
    distribution: Dict[str, int] = {}
    for d in by_driver.values():
        for k, v in d.items():
            if k.startswith("CDF_L1_"):
                distribution[k] = distribution.get(k, 0) + int(v)

    summary = dict(meta)
    summary["cdfSummary"] = {
        "totalDrivers": len(by_driver),
        "totalEvents": len(events),
        "totalNegative": sum(
            int(d["CDF_NegativeFeedback"]) for d in by_driver.values()
        ),
    }
    summary["cdfL1Distribution"] = distribution

    drivers = list(by_driver.values())
    drivers.sort(
        key=lambda d: int(d.get("CDF_NegativeFeedback") or 0), reverse=True
    )
    return summary, drivers


# ===============================
# DWC / IADC (weekly HTML report)
# ===============================
def _is_dwc_html(filename: str, content: bytes) -> bool:
    name = (filename or "").lower()
    if ("dwc" in name or "iadc" in name) and (
        name.endswith(".html") or name.endswith(".htm")
    ):
        return True
    # Sniff: Amazon's title is "… DWC/IADC Report …".
    head = content[:4096].decode("utf-8", errors="ignore").lower()
    return ("dwc" in head and "iadc" in head) and (
        "<html" in head or "<table" in head
    )


def _parse_dwc_meta_from_filename(filename: str) -> Dict[str, Any]:
    """e.g. "DE-AION-DBY5-DWC-IADC-Report_2026-26.html"
    → {country:'DE', dsp:'AION', stationCode:'DBY5', year:2026, weekNumber:26}.
    """
    out: Dict[str, Any] = {}
    if not filename:
        return out
    base = filename.rsplit("/", 1)[-1]
    m = re.search(r"([A-Z]{2})-([A-Z]+)-([A-Z0-9]+)-DWC", base, re.I)
    if m:
        out["country"] = m.group(1).upper()
        out["dsp"] = m.group(2).upper()
        out["stationCode"] = m.group(3).upper()
    w = re.search(r"(20\d{2})[-_]?W?(\d{1,2})(?!\d)", base)
    if w:
        out["year"] = int(w.group(1))
        out["weekNumber"] = int(w.group(2))
    return out


def _dwc_percent(token: Optional[str]) -> Optional[float]:
    if token is None:
        return None
    t = token.strip().replace("%", "").replace(",", ".")
    if not t:
        return None
    try:
        return round(float(t), 2)
    except ValueError:
        return None


def _dwc_weekly_bucket_totals(html: str):
    """Reads the weekly aggregate "Bucket / Reason / Total" tables and returns
    `{ total: bucketName }`. Those tables carry clean, localized bucket names
    (via `data-expand-row-id`) plus each bucket's station-wide weekly total —
    which we use to self-calibrate which column of the wide per-driver table
    is which bucket (see `_extract_dwc_from_html`)."""
    from bs4 import BeautifulSoup  # type: ignore

    # Restrict to the weekly section (before the first "Day:" heading) so the
    # daily aggregates don't shadow the weekly totals.
    wi = html.find("Week:")
    di = html.find("Day:", wi if wi >= 0 else 0)
    weekly = html[wi:di] if wi >= 0 and di > wi else html
    wsoup = BeautifulSoup(weekly, "html.parser")

    by_total: Dict[int, str] = {}
    for th in wsoup.find_all(attrs={"data-expand-row-id": True}):
        name = th.get("data-expand-row-id")
        tr = th.find_parent("tr")
        if tr is None:
            continue
        nums = [
            (c.get_text() or "").strip()
            for c in tr.find_all(["th", "td"])
            if (c.get_text() or "").strip().isdigit()
        ]
        if nums:
            total = int(nums[-1])
            by_total.setdefault(total, name)  # first (weekly) wins
    return by_total


def _extract_dwc_from_html(content: bytes, filename: str):
    """Parse the weekly "DWC/IADC Misses by Transporter ID" table into
    per-driver DWC% / IADC% rows + per-driver miss counts by category, plus
    the week-level DWC% / IADC%."""
    try:
        from bs4 import BeautifulSoup  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "DWC HTML parsing requires beautifulsoup4 — add it to "
            "requirements.txt and redeploy."
        ) from e

    meta = _parse_dwc_meta_from_filename(filename)
    html = content.decode("utf-8", errors="replace")
    soup = BeautifulSoup(html, "html.parser")

    # Week-level DWC% / IADC%. The weekly section is rendered before the
    # daily ones, so the first occurrence is the weekly value.
    text_all = soup.get_text("\n")
    mdwc = re.search(r"DWC\s*%\s*:\s*([\d.,]+)\s*%", text_all)
    miadc = re.search(r"IADC\s*%\s*:\s*([\d.,]+)\s*%", text_all)
    week_dwc = _dwc_percent(mdwc.group(1)) if mdwc else None
    week_iadc = _dwc_percent(miadc.group(1)) if miadc else None

    # Locate the per-transporter table + remember its header/group rows.
    target_table = None
    header_labels: List[str] = []
    group_row = None
    for table in soup.find_all("table"):
        trs = table.find_all("tr")
        for tr in trs:
            cells = tr.find_all(["th", "td"], recursive=False)
            labels = [(c.get_text() or "").strip() for c in cells]
            lower = [x.lower() for x in labels]
            if "transporter id" in lower and "dwc %" in lower:
                header_labels = labels
                target_table = table
                group_row = trs[0] if trs else None
                break
        if target_table is not None:
            break

    if target_table is None or not header_labels:
        summary = dict(meta)
        summary["dwcSummary"] = {
            "dwcPct": week_dwc, "iadcPct": week_iadc, "totalDrivers": 0,
        }
        return summary, []

    lower = [x.lower() for x in header_labels]
    tid_i = lower.index("transporter id") if "transporter id" in lower else None
    dwc_i = lower.index("dwc %") if "dwc %" in lower else None
    iadc_i = lower.index("iadc %") if "iadc %" in lower else None
    total_cols = [i for i, x in enumerate(lower) if x == "total"]

    # Group (ROW0) label per leaf column — used to label any miss column whose
    # total can't be matched to an aggregate bucket.
    leaf = len(header_labels)
    group_of: List[str] = [""] * leaf
    if group_row is not None:
        ci = 0
        for c in group_row.find_all(["th", "td"], recursive=False):
            span = int(c.get("colspan", 1))
            g = (c.get_text() or "").strip()
            for _ in range(span):
                if ci < leaf:
                    group_of[ci] = g
                    ci += 1

    def looks_like_tid(x: str) -> bool:
        return bool(re.match(r"^A[A-Z0-9]{10,}$", (x or "").strip().upper()))

    # Pass 1: collect each driver's raw values at every "Total" column, and
    # accumulate the column sums across drivers.
    rows_raw: List[Dict[str, Any]] = []
    col_sum: Dict[int, int] = {i: 0 for i in total_cols}
    for tr in target_table.find_all("tr"):
        # The Transporter ID sits in a row-header <th>; include it so the
        # column positions line up with the header row.
        tds = tr.find_all(["th", "td"], recursive=False)
        if not tds or tid_i is None or tid_i >= len(tds):
            continue
        tid = (tds[tid_i].get_text() or "").strip()
        if not looks_like_tid(tid):
            continue

        def cell(i: Optional[int]) -> str:
            if i is None or i >= len(tds):
                return ""
            return (tds[i].get_text() or "").strip()

        counts: Dict[int, int] = {}
        for i in total_cols:
            txt = cell(i)
            n = int(txt) if txt.isdigit() else 0
            counts[i] = n
            col_sum[i] += n
        rows_raw.append({
            "tid": tid,
            "dwc": _dwc_percent(cell(dwc_i)),
            "iadc": _dwc_percent(cell(iadc_i)),
            "counts": counts,
        })

    # Self-calibrate: match each Total column to a bucket by its column-sum ==
    # the weekly aggregate bucket total. Unmatched columns fall back to their
    # ROW0 group name so nothing is silently mislabelled.
    bucket_totals = _dwc_weekly_bucket_totals(html)
    col_name: Dict[int, str] = {}
    used = set()
    for i in total_cols:
        name = bucket_totals.get(col_sum[i])
        if name and name not in used:
            col_name[i] = name
            used.add(name)
        else:
            g = group_of[i] or "Misses"
            col_name[i] = f"{g} (Total)"

    by_driver: List[Dict[str, Any]] = []
    for r in rows_raw:
        misses: Dict[str, int] = {}
        for i in total_cols:
            n = r["counts"].get(i, 0)
            if n > 0:
                misses[col_name[i]] = misses.get(col_name[i], 0) + n
        by_driver.append({
            "Transporter ID": r["tid"],
            "DWC_DwcPct": r["dwc"],
            "DWC_IadcPct": r["iadc"],
            "DWC_Misses": misses,
        })

    summary = dict(meta)
    summary["dwcSummary"] = {
        "dwcPct": week_dwc,
        "iadcPct": week_iadc,
        "totalDrivers": len(by_driver),
    }
    # Worst IADC first — those drivers need the most attention.
    by_driver.sort(
        key=lambda d: (d.get("DWC_IadcPct") is None, d.get("DWC_IadcPct") or 0.0)
    )
    return summary, by_driver


# ===============================
# AMAZON DSP INVOICE / GUTSCHRIFT
# ===============================
def _is_amazon_dsp_invoice_pdf(pdf: "pdfplumber.PDF", filename: str) -> bool:
    """Detect an Amazon DSP weekly credit note (Gutschrift / Rechnung)."""
    name = (filename or "").lower()
    if any(tag in name for tag in ("rechnung", "gutschrift", "invoice")):
        return True
    text = ""
    try:
        text = clean_str(pdf.pages[0].extract_text() or "").lower()
    except Exception:
        text = ""
    signatures = (
        "details des rechnungsdatums",
        "rechnung - inv-",
        "gutschrift",
        "für zustellungen in woche",
        "fur zustellungen in woche",
    )
    return any(sig in text for sig in signatures)


def _us_date_to_iso(raw: str) -> Optional[str]:
    """'06/01/2026' (MM/DD/YYYY) -> '2026-06-01'."""
    s = clean_str(raw)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if not m:
        return None
    mm, dd, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not (1 <= mm <= 12 and 1 <= dd <= 31):
        return None
    return f"{yyyy:04d}-{mm:02d}-{dd:02d}"


def _extract_invoice(pdf: "pdfplumber.PDF", filename: str) -> Dict[str, Any]:
    """Parse the weekly DSP credit note into header + per-day line items.

    The detail table ("Details des Rechnungsdatums") has columns:
    Datum (MM/DD/YYYY) · Beschreibung · Stückpreis · Menge · Summe.
    We collect every row across all pages/tables whose first cell is a
    US-format date — that naturally excludes the aggregated top summary
    table (its first column is text, not a date).
    """
    full_text = " ".join(
        clean_str(p.extract_text() or "") for p in pdf.pages
    )

    res: Dict[str, Any] = {
        "invoiceNumber": None,
        "week": None,
        "rangeStart": None,
        "rangeEnd": None,
        "total": None,
        "lines": [],
    }

    if m := re.search(r"(INV-[A-Z0-9]+-\d+)", full_text, re.IGNORECASE):
        res["invoiceNumber"] = m.group(1)

    # "Für Zustellungen in Woche 23: 05/31/2026 to 06/06/2026"
    if m := re.search(
        r"[Ww]oche\s*(\d{1,2})\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})\s*"
        r"(?:to|bis|-|–|—)\s*(\d{1,2}/\d{1,2}/\d{4})",
        full_text,
    ):
        res["week"] = int(m.group(1))
        res["rangeStart"] = _us_date_to_iso(m.group(2))
        res["rangeEnd"] = _us_date_to_iso(m.group(3))
    elif m := re.search(r"[Ww]oche\s*(\d{1,2})", full_text):
        res["week"] = int(m.group(1))

    if m := re.search(r"Gesamtbetrag\s*([0-9.,]+)", full_text):
        res["total"] = to_num(m.group(1))

    lines: List[Dict[str, Any]] = []
    seen = set()
    for page in pdf.pages:
        for table in page.extract_tables() or []:
            if not table:
                continue
            for row in table:
                if not row:
                    continue
                cells = [clean_str(c) for c in row]
                iso = _us_date_to_iso(cells[0]) if cells else None
                if not iso or len(cells) < 5:
                    continue
                desc = cells[1]
                unit = to_num(cells[2])
                qty = to_num(cells[3])
                total = to_num(cells[4])
                if not desc:
                    continue
                key = (iso, desc, qty, total)
                if key in seen:
                    continue
                seen.add(key)
                lines.append({
                    "date": iso,
                    "serviceType": desc,
                    "unitPrice": unit,
                    "quantity": qty,
                    "total": total,
                })

    # Fallback: if the detail table has no ruling lines, pdfplumber won't
    # return it via extract_tables(). Parse the per-day rows from raw text
    # instead: "06/01/2026  <description>  158,25 €  2.0  316,50 €".
    if not lines:
        row_re = re.compile(
            r"(\d{1,2}/\d{1,2}/\d{4})\s+(.+?)\s+"
            r"([0-9.,]+)\s*€\s+([0-9.,]+)\s+([0-9.,]+)\s*€"
        )
        for page in pdf.pages:
            for raw_line in (page.extract_text() or "").split("\n"):
                ln = clean_str(raw_line)
                m = row_re.search(ln)
                if not m:
                    continue
                iso = _us_date_to_iso(m.group(1))
                if not iso:
                    continue
                desc = clean_str(m.group(2))
                qty = to_num(m.group(4))
                total = to_num(m.group(5))
                key = (iso, desc, qty, total)
                if not desc or key in seen:
                    continue
                seen.add(key)
                lines.append({
                    "date": iso,
                    "serviceType": desc,
                    "unitPrice": to_num(m.group(3)),
                    "quantity": qty,
                    "total": total,
                })

    lines.sort(key=lambda r: (r["date"], r["serviceType"]))
    res["lines"] = lines
    return res


@app.get("/health")
def health():
    return {"ok": True}

@app.post("/parse", response_model=ParserResponse)
async def parse_pdf(file: UploadFile = File(...)):
    content = await file.read()
    filename = file.filename or ""

    # HTML route: DWC / IADC weekly report (per-driver DWC% / IADC%)
    if _is_dwc_html(filename, content):
        summary, drivers = _extract_dwc_from_html(content, filename)
        return {
            "count": len(drivers),
            "drivers": drivers,
            "summary": summary or None,
        }

    # HTML route: CDF (Customer Delivery Feedback) weekly report
    if _is_cdf_html(filename, content):
        summary, drivers = _extract_cdf_from_html(content, filename)
        return {
            "count": len(drivers),
            "drivers": drivers,
            "summary": summary or None,
        }

    # XLSX route: Cortex VehiclesData export (DSP fleet from Amazon)
    if _is_vehicles_xlsx(filename, content):
        summary, vehicles = _extract_vehicles_from_xlsx(content, filename)
        return {
            "count": len(vehicles),
            "drivers": [],
            "vehicles": vehicles,
            "summary": summary or None,
        }

    # XLSX route: Concessions report (legacy OR new DSC format)
    if _is_concessions_xlsx(filename):
        # Peek at the workbook sheets to choose the right parser. The new
        # 2026 "DSC Concession" XLSX has a different structure with
        # event-level rows + per-week date columns.
        from openpyxl import load_workbook as _peek_wb
        try:
            peek = _peek_wb(io.BytesIO(content), data_only=True, read_only=True)
            is_dsc = _is_dsc_concessions_xlsx_by_sheets(peek)
        except Exception:
            is_dsc = False
        # filename hint: "DSC" anywhere in name → also assume DSC format
        if "dsc" in (filename or "").lower():
            is_dsc = True

        if is_dsc:
            summary, drivers = _extract_dsc_concessions_from_xlsx(
                content, filename
            )
        else:
            summary, drivers = _extract_concessions_from_xlsx(content, filename)
        return {
            "count": len(drivers),
            "drivers": drivers,
            "summary": summary or None,
        }

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        # Amazon DSP weekly credit note (Gutschrift) — per-day, per-service
        # line items for the Payment Check reconciliation.
        if _is_amazon_dsp_invoice_pdf(pdf, filename):
            invoice = _extract_invoice(pdf, filename)
            return {
                "count": len(invoice.get("lines", [])),
                "drivers": [],
                "invoice": invoice,
            }
        if _is_pod_quality_pdf(pdf, filename):
            summary = _extract_pod_quality_summary(pdf, filename)
            drivers = _extract_pod_quality_drivers(pdf)
        else:
            summary = extract_summary(pdf)
            week_number = summary.get("weekNumber") if summary else None
            year = summary.get("year") if summary else None  # <-- ADDED

            # ONLY CHANGE: pass year through so FinalScore uses (year, week)
            drivers = extract_driver_rows(pdf, week_number=week_number, year=year)

            add_ranking_and_status(drivers, week_number=week_number, year=year)

    for d in drivers:
        if isinstance(d, dict) and "_cdf_mode" in d:
            d.pop("_cdf_mode", None)

    return {
        "count": len(drivers),
        "drivers": drivers,
        "summary": summary or None
    }
